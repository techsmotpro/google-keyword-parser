import requests
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
API_KEY = os.getenv('SERPAPI_KEY')
if not API_KEY:
    raise RuntimeError("SERPAPI_KEY environment variable is required")


def check_ranking(keyword, domain):
    found = False
    all_results = []

    for page in range(10):
        start = page * 10

        encoded_keyword = urllib.parse.quote(keyword)
        url = f"https://serpapi.com/search.json?q={encoded_keyword}&start={start}&api_key={API_KEY}"

        try:
            response = requests.get(url)
            response.raise_for_status()  # Raise exception for HTTP errors
            data = response.json()

            results = data.get("organic_results", [])
            
            # Collect all results for debugging
            for i, result in enumerate(results):
                position = start + i + 1
                link = result.get("link", "")
                title = result.get("title", "")
                all_results.append({
                    'position': position,
                    'page': page + 1,
                    'title': title,
                    'link': link
                })

                if domain in link:
                    print("✅ FOUND!")
                    print(f"Keyword: {keyword}")
                    print(f"Page: {page + 1}")
                    print(f"Position: {position}")
                    print(f"Title: {title}")
                    print(f"Link: {link}")
                    found = True
                    return

        except requests.exceptions.RequestException as err:
            print(f"Error fetching page {page + 1}: {err}")
        except Exception as err:
            print(f"Unexpected error on page {page + 1}: {err}")

    if not found:
        print("❌ Not found in top 100")
        print("\n📊 All Results Found (Top 100):")
        print("-" * 80)
        print(f"{'Position':<8} {'Page':<5} {'Title':<50} {'Domain'}")
        print("-" * 80)
        for result in all_results:
            # Extract domain from URL
            try:
                result_domain = result['link'].split('//')[1].split('/')[0]
            except:
                result_domain = result['link']
            
            # Truncate title if too long
            truncated_title = (result['title'][:47] + '...') if len(result['title']) > 47 else result['title']
            
            print(f"{result['position']:<8} {result['page']:<5} {truncated_title:<50} {result_domain}")


# TEST with multiple keywords, detailed results, and Excel export
if __name__ == "__main__":
    keywords = [
        # Bangalore core
        "visa services in bangalore",
        "visa consultant in bangalore",
        
        # Slightly easier variations
        "tourist visa services bangalore",
        "student visa consultants bangalore",
        "business visa services bangalore",

        # Country + city
        "us visa services bangalore",
        "uk visa services bangalore",
        "canada visa services bangalore",
        "australia visa services bangalore",
        "schengen visa services bangalore",

        # Long tail
        "how to apply us visa from bangalore",
        "documents required for canada visa india",
        "uk visa processing time india",
        "schengen visa agents india",
        "uae tourist visa from india",

        # Location-based
        "visa services cv raman nagar bangalore",
        "visa consultant whitefield bangalore",
        "visa services marathahalli bangalore",

        # Service intent keywords
        "apply tourist visa online india",
        "visa assistance services india",
        "schengen visa agent",
        "uk visa agent",
        "us visa agent",
        "usa visa agent",
        "singapore visa agent",
        "dubai visa agent"
    ]
    
    target_domain = "smotvisa.com"
    print(f"🔍 Searching for domain '{target_domain}' in Google results\n")
    
    found_keywords = []
    not_found_keywords = []
    all_search_results = []
    
    for keyword in keywords:
        print(f"{'='*80}")
        print(f"Checking: '{keyword}'")
        print(f"{'='*80}")
        
        found = False
        keyword_results = []
        
        # Search up to page 11 (110 results)
        for page in range(11):
            start = page * 10
            
            encoded_keyword = urllib.parse.quote(keyword)
            url = f"https://serpapi.com/search.json?q={encoded_keyword}&start={start}&api_key={API_KEY}"
            
            try:
                response = requests.get(url)
                response.raise_for_status()
                data = response.json()
                
                results = data.get("organic_results", [])
                
                for i, result in enumerate(results):
                    position = start + i + 1
                    link = result.get("link", "")
                    title = result.get("title", "")
                    
                    # Extract domain from URL
                    try:
                        domain = link.split('//')[1].split('/')[0]
                    except:
                        domain = link
                    
                    keyword_results.append({
                        'keyword': keyword,
                        'position': position,
                        'page': page + 1,
                        'title': title,
                        'link': link,
                        'domain': domain
                    })
                    
                    if target_domain in link:
                        print(f"✅ FOUND!")
                        print(f"Keyword: {keyword}")
                        print(f"Page: {page + 1}")
                        print(f"Position: {position}")
                        print(f"Title: {title}")
                        print(f"Link: {link}")
                        print()
                        
                        found_keywords.append({
                            'keyword': keyword,
                            'page': page + 1,
                            'position': position,
                            'title': title,
                            'link': link
                        })
                        found = True
                        break
                
                if found:
                    break
                    
            except requests.exceptions.RequestException as err:
                print(f"Error fetching page {page + 1}: {err}")
            except Exception as err:
                print(f"Unexpected error on page {page + 1}: {err}")
        
        all_search_results.extend(keyword_results)
        
        if not found:
            print("❌ Not found in top 110 results")
            print("\n📊 All Results Found (Top 110):")
            print("-" * 80)
            print(f"{'Position':<8} {'Page':<5} {'Title':<50} {'Domain'}")
            print("-" * 80)
            
            for result in keyword_results:
                # Truncate title if too long
                truncated_title = (result['title'][:47] + '...') if len(result['title']) > 47 else result['title']
                print(f"{result['position']:<8} {result['page']:<5} {truncated_title:<50} {result['domain']}")
            
            print()
            not_found_keywords.append(keyword)
    
    # Summary report
    print("\n" + "="*80)
    print(f"📊 Summary for domain '{target_domain}'")
    print("="*80)
    
    print(f"\n✅ Keywords Found: {len(found_keywords)}")
    if found_keywords:
        print("-"*60)
        for item in found_keywords:
            print(f"{item['keyword']:<40} | Page {item['page']:<2} | Position {item['position']:<3}")
    
    print(f"\n❌ Keywords Not Found: {len(not_found_keywords)}")
    if not_found_keywords:
        print("-"*60)
        for keyword in not_found_keywords:
            print(f"{keyword}")
    
    # Export results to Excel
    print("\n📊 Exporting results to Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Rankings"
    
    # Define columns
    columns = [
        "Keyword",
        "Position",
        "Page",
        "Title",
        "Domain",
        "Link",
        "Found Target Domain"
    ]
    
    # Write header
    for col, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=column_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Write data
    row = 2
    for result in all_search_results:
        ws.cell(row=row, column=1, value=result['keyword'])
        ws.cell(row=row, column=2, value=result['position'])
        ws.cell(row=row, column=3, value=result['page'])
        ws.cell(row=row, column=4, value=result['title'])
        ws.cell(row=row, column=5, value=result['domain'])
        ws.cell(row=row, column=6, value=result['link'])
        ws.cell(row=row, column=7, value="Yes" if target_domain in result['domain'] else "No")
        
        # Apply borders to all cells
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Auto-fit columns
    for col in range(1, len(columns) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    summary_columns = [
        "Keyword",
        "Status",
        "Page",
        "Position",
        "Title",
        "Link"
    ]
    
    # Write summary header
    for col, column_name in enumerate(summary_columns, 1):
        cell = ws_summary.cell(row=1, column=col, value=column_name)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Write found keywords
    row = 2
    for item in found_keywords:
        ws_summary.cell(row=row, column=1, value=item['keyword'])
        ws_summary.cell(row=row, column=2, value="Found")
        ws_summary.cell(row=row, column=3, value=item['page'])
        ws_summary.cell(row=row, column=4, value=item['position'])
        ws_summary.cell(row=row, column=5, value=item['title'])
        ws_summary.cell(row=row, column=6, value=item['link'])
        
        for col in range(1, len(summary_columns) + 1):
            ws_summary.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Write not found keywords
    for keyword in not_found_keywords:
        ws_summary.cell(row=row, column=1, value=keyword)
        ws_summary.cell(row=row, column=2, value="Not Found")
        
        for col in range(1, len(summary_columns) + 1):
            ws_summary.cell(row=row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Auto-fit summary columns
    for col in range(1, len(summary_columns) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for cell in ws_summary[column_letter]:
            try:
                cell_value = str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save the Excel file
    excel_filename = "keyword_rankings.xlsx"
    wb.save(excel_filename)
    print(f"✅ Excel file saved successfully: {excel_filename}")
    
    # Print Excel file location
    import os
    file_path = os.path.abspath(excel_filename)
    print(f"📁 File location: {file_path}")

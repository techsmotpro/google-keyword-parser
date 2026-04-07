from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import requests
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import tempfile

app = Flask(__name__, instance_path=sys.path[0])
CORS(app)

# Get API key from environment variable
API_KEY = os.getenv('SERPAPI_KEY')
# Handle case where SERPAPI_KEY is not set
if not API_KEY:
    # Log error message instead of raising exception to avoid crashing the app
    import logging
    logging.error("SERPAPI_KEY environment variable is not set")
    API_KEY = None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        results = data['results']
        domain = data['domain']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Keyword Rankings"
        
        columns = [
            "Keyword",
            "Position",
            "Page",
            "Title",
            "Domain",
            "Link",
            "Found Target Domain"
        ]
        
        for col, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row = 2
        for result in results:
            for item in result['all_results']:
                ws.cell(row=row, column=1, value=result['keyword'])
                ws.cell(row=row, column=2, value=item['position'])
                ws.cell(row=row, column=3, value=item['page'])
                ws.cell(row=row, column=4, value=item['title'])
                ws.cell(row=row, column=5, value=item['domain'])
                ws.cell(row=row, column=6, value=item['link'])
                ws.cell(row=row, column=7, value="Yes" if domain in item['domain'] else "No")
                
                for col in range(1, len(columns) + 1):
                    ws.cell(row=row, column=col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                row += 1
        
        for col in range(1, len(columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws_summary = wb.create_sheet(title="Summary")
        summary_columns = [
            "Keyword",
            "Status",
            "Page",
            "Position",
            "Title",
            "Link"
        ]
        
        for col, column_name in enumerate(summary_columns, 1):
            cell = ws_summary.cell(row=1, column=col, value=column_name)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row = 2
        for result in results:
            ws_summary.cell(row=row, column=1, value=result['keyword'])
            ws_summary.cell(row=row, column=2, value="Found" if result['status'] == 'found' else "Not Found")
            ws_summary.cell(row=row, column=3, value=result['page'])
            ws_summary.cell(row=row, column=4, value=result['position'])
            ws_summary.cell(row=row, column=5, value=result['title'])
            ws_summary.cell(row=row, column=6, value=result['link'])
            
            for col in range(1, len(summary_columns) + 1):
                ws_summary.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            row += 1
        
        for col in range(1, len(summary_columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for cell in ws_summary[column_letter]:
                try:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Use temp file in /tmp directory (writable on Vercel)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir='/tmp')
        temp_file.close()
        wb.save(temp_file.name)
        
        return send_file(temp_file.name, as_attachment=True, download_name='keyword_rankings.xlsx')
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

@app.route('/check-ranking', methods=['POST'])
def check_ranking():
    try:
        data = request.json
        keywords = data['keywords']
        domain = data['domain']
        
        # Check if API key is available
        if not API_KEY:
            return jsonify({
                'success': False,
                'error': 'SERPAPI_KEY environment variable is not configured. Please set it in the Vercel dashboard.'
            }), 500
        
        results = []
        
        for keyword in keywords:
            keyword = keyword.strip()
            if not keyword:
                continue
                
            found = False
            keyword_results = []
            
            for page in range(10):
                start = page * 10
                
                encoded_keyword = urllib.parse.quote(keyword)
                url = f"https://serpapi.com/search.json?q={encoded_keyword}&start={start}&api_key={API_KEY}"
                
                try:
                    response = requests.get(url, timeout=10)
                    response.raise_for_status()
                    data = response.json()
                    
                    organic_results = data.get("organic_results", [])
                    
                    for i, result in enumerate(organic_results):
                        position = start + i + 1
                        link = result.get("link", "")
                        title = result.get("title", "")
                        
                        try:
                            result_domain = link.split('//')[1].split('/')[0]
                        except:
                            result_domain = link
                        
                        keyword_results.append({
                            'position': position,
                            'page': page + 1,
                            'title': title,
                            'link': link,
                            'domain': result_domain
                        })
                        
                        if domain in link:
                            found = True
                            results.append({
                                'keyword': keyword,
                                'status': 'found',
                                'page': page + 1,
                                'position': position,
                                'title': title,
                                'link': link,
                                'all_results': keyword_results
                            })
                            break
                    
                    if found:
                        break
                        
                except requests.exceptions.RequestException as err:
                    print(f"Error fetching page {page + 1}: {err}")
                    continue
                except Exception as err:
                    print(f"Unexpected error on page {page + 1}: {err}")
                    continue
            
            if not found:
                results.append({
                    'keyword': keyword,
                    'status': 'not_found',
                    'page': None,
                    'position': None,
                    'title': None,
                    'link': None,
                    'all_results': keyword_results
                })
        
        return jsonify({
            'success': True,
            'results': results,
            'domain': domain
        })
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        })

if __name__ == '__main__':
    app.run(debug=True, port=8080)
